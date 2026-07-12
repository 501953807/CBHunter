"""Product import and export helpers for CSV/XLSX files."""

import csv
import io
import json
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.product import Product
from app.schemas.product import ProductCreate
from app.services.product_service import create_product, validate_product_name_quality

PRODUCT_COLUMNS = [
    ("sku", "SKU"),
    ("name", "商品名称"),
    ("description", "商品描述"),
    ("brand", "品牌"),
    ("category_id", "品类ID"),
    ("cost_price", "成本价"),
    ("weight_g", "重量g"),
    ("status", "状态"),
    ("notes", "备注"),
    ("tags", "标签"),
    ("images", "图片"),
    ("attributes", "属性JSON"),
    ("dimensions", "尺寸JSON"),
]


async def export_products(db: AsyncSession, user_id: str) -> list[Product]:
    result = await db.execute(
        select(Product).where(Product.user_id == user_id).order_by(Product.updated_at.desc())
    )
    return list(result.scalars().all())


def build_product_csv(products: list[Product]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([label for _, label in PRODUCT_COLUMNS])
    for product in products:
        writer.writerow(_product_row(product))
    return ("\ufeff" + buffer.getvalue()).encode("utf-8")


def build_product_xlsx(products: list[Product]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet.append([label for _, label in PRODUCT_COLUMNS])
    for product in products:
        sheet.append(_product_row(product))
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


async def import_products_from_upload(db: AsyncSession, user_id: str, filename: str, content: bytes) -> dict:
    rows = _read_rows(filename, content)
    created: list[Product] = []
    errors: list[dict] = []
    for index, row in enumerate(rows, start=2):
        try:
            req = _row_to_product(row)
        except ValueError as exc:
            errors.append({"row": index, "error": str(exc)})
            continue
        product = await create_product(db, user_id, req)
        created.append(product)
    return {
        "created_count": len(created),
        "failed_count": len(errors),
        "errors": errors,
        "product_ids": [item.id for item in created],
    }


def _product_row(product: Product) -> list[Any]:
    return [
        product.sku,
        product.name,
        product.description or "",
        product.brand or "",
        product.category_id or "",
        product.cost_price if product.cost_price is not None else "",
        product.weight_g if product.weight_g is not None else "",
        product.status,
        product.notes or "",
        ";".join(product.tags or []),
        ";".join(product.images or []),
        json.dumps(product.attributes or {}, ensure_ascii=False),
        json.dumps(product.dimensions or {}, ensure_ascii=False),
    ]


def _read_rows(filename: str, content: bytes) -> list[dict[str, Any]]:
    lower_name = filename.lower()
    if lower_name.endswith(".xlsx"):
        return _read_xlsx_rows(content)
    if lower_name.endswith(".csv"):
        return _read_csv_rows(content)
    raise ValueError("仅支持 CSV 或 XLSX 商品文件")


def _read_csv_rows(content: bytes) -> list[dict[str, Any]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("CSV 文件必须使用 UTF-8 编码") from exc
    return list(csv.DictReader(io.StringIO(text)))


def _read_xlsx_rows(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    result = []
    for row in rows[1:]:
        result.append({headers[index]: value for index, value in enumerate(row) if index < len(headers)})
    return result


def _row_to_product(row: dict[str, Any]) -> ProductCreate:
    values = {_normalize_key(key): value for key, value in row.items()}
    name = _text(values.get("name"))
    if not name:
        raise ValueError("商品名称不能为空")
    validate_product_name_quality(name)
    return ProductCreate(
        sku=_optional_text(values.get("sku")),
        name=name,
        description=_optional_text(values.get("description")),
        brand=_optional_text(values.get("brand")),
        category_id=_optional_text(values.get("category_id")),
        cost_price=_optional_float(values.get("cost_price")),
        weight_g=_optional_float(values.get("weight_g")),
        status=_optional_text(values.get("status")) or "draft",
        notes=_optional_text(values.get("notes")),
        tags=_split_list(values.get("tags")),
        images=_split_list(values.get("images")),
        attributes=_json_dict(values.get("attributes")),
        dimensions=_json_dict(values.get("dimensions")),
    )


def _normalize_key(key: str) -> str:
    mapping = {field: field for field, _ in PRODUCT_COLUMNS}
    mapping.update({label: field for field, label in PRODUCT_COLUMNS})
    mapping.update({"商品名": "name", "成本": "cost_price", "重量": "weight_g", "品类": "category_id"})
    return mapping.get(str(key).strip(), str(key).strip())


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _optional_text(value: Any) -> str | None:
    text = _text(value)
    return text or None


def _optional_float(value: Any) -> float | None:
    text = _text(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError as exc:
        raise ValueError(f"数值字段格式错误: {text}") from exc


def _split_list(value: Any) -> list[str]:
    text = _text(value)
    if not text:
        return []
    return [item.strip() for item in text.replace(",", ";").split(";") if item.strip()]


def _json_dict(value: Any) -> dict:
    text = _text(value)
    if not text:
        return {}
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError as exc:
        raise ValueError(f"JSON 字段格式错误: {text}") from exc
    if not isinstance(parsed, dict):
        raise ValueError("JSON 字段必须是对象")
    return parsed
